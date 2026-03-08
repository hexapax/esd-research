#!/usr/bin/env python3
"""Export esd-dataset/ YAML frontmatter to CSV and XLSX files."""

import yaml
import csv
import os
import glob
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DATASET_DIR = "/opt/repos/esd-research/esd-dataset"
OUTPUT_DIR = "/opt/repos/esd-research/esd-dataset/exports"
DATE_STAMP = datetime.now().strftime("%Y%m%d")

def parse_yaml_frontmatter(filepath):
    """Extract YAML frontmatter from a markdown file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Match YAML between --- markers
    match = re.match(r'^---\s*\n(.*?)\n---', content, re.DOTALL)
    if not match:
        return None

    try:
        data = yaml.safe_load(match.group(1))
        return data
    except yaml.YAMLError as e:
        print(f"YAML error in {filepath}: {e}")
        return None

def safe_get(d, *keys, default=""):
    """Safely navigate nested dicts."""
    for k in keys:
        if isinstance(d, dict):
            d = d.get(k, default)
        else:
            return default
    return d if d is not None else default

def get_first_source_url(data):
    """Get the first source URL from sources list."""
    sources = data.get('sources', [])
    if sources and isinstance(sources, list) and len(sources) > 0:
        first = sources[0]
        if isinstance(first, dict):
            return first.get('url', '')
    return ''

def get_all_source_urls(data):
    """Get all source URLs joined by newline."""
    sources = data.get('sources', [])
    urls = []
    if sources and isinstance(sources, list):
        for s in sources:
            if isinstance(s, dict) and s.get('url'):
                urls.append(s['url'])
    return "\n".join(urls)

def get_all_sources_detail(data):
    """Get all sources with type and outlet."""
    sources = data.get('sources', [])
    details = []
    if sources and isinstance(sources, list):
        for s in sources:
            if isinstance(s, dict):
                parts = []
                if s.get('outlet'): parts.append(s['outlet'])
                if s.get('type'): parts.append(f"({s['type']})")
                if s.get('url'): parts.append(s['url'])
                if s.get('date'): parts.append(f"[{s['date']}]")
                details.append(" ".join(parts))
    return "\n".join(details)

def victims_summary(data):
    """Create a summary string of all victims."""
    victims = data.get('victims', [])
    if not victims or not isinstance(victims, list):
        return ""
    parts = []
    for v in victims:
        if not isinstance(v, dict):
            continue
        first = v.get('first_name', 'Unknown')
        last = v.get('last_name', 'Unknown')
        age = v.get('age', '')
        gender = v.get('gender', '')
        name = f"{first} {last}".strip()
        if name == "Unknown Unknown":
            name = "Unknown"
        suffix = []
        if age: suffix.append(str(age))
        if gender: suffix.append(gender)
        if suffix:
            name += f" ({', '.join(suffix)})"
        parts.append(name)
    return "; ".join(parts)

def list_to_str(val):
    """Convert a list to semicolon-separated string."""
    if isinstance(val, list):
        return "; ".join(str(x) for x in val)
    return str(val) if val else ""

def build_incident_row_small(data):
    """Build a row with Nicole's requested columns."""
    victims = data.get('victims', []) or []
    # For incident-level, concatenate victim names
    first_victim = victims[0] if victims and isinstance(victims, list) and len(victims) > 0 else {}

    return {
        'Incident ID': safe_get(data, 'incident_id'),
        'Date': safe_get(data, 'date'),
        'Year': safe_get(data, 'year'),
        'First Name': safe_get(first_victim, 'first_name') if isinstance(first_victim, dict) else '',
        'Last Name': safe_get(first_victim, 'last_name') if isinstance(first_victim, dict) else '',
        'State': safe_get(data, 'state'),
        'County': safe_get(data, 'county'),
        'City': safe_get(data, 'city'),
        'Age': safe_get(first_victim, 'age') if isinstance(first_victim, dict) else '',
        'Gender': safe_get(first_victim, 'gender') if isinstance(first_victim, dict) else '',
        'Race': safe_get(first_victim, 'race') if isinstance(first_victim, dict) else '',
        'Fatalities': safe_get(data, 'fatality_count', default=0),
        'Injuries': safe_get(data, 'injury_count', default=0),
        'Near Misses': safe_get(data, 'near_miss_count', default=0),
        'Body of Water': safe_get(data, 'body_of_water'),
        'Facility Name': safe_get(data, 'facility_name'),
        'Notes': safe_get(data, 'notes'),
        'Source URL': get_first_source_url(data),
    }

def build_incident_row_medium(data):
    """Nicole's columns + research metadata."""
    row = build_incident_row_small(data)
    row.update({
        'All Victims': victims_summary(data),
        'Incident Type': safe_get(data, 'incident_type'),
        'Verification Level': safe_get(data, 'verification_level'),
        'ESDPA Listed': safe_get(data, 'esdpa_listed'),
        'ESDPA Date Correct': safe_get(data, 'esdpa_date_correct'),
        'Discovery Source': safe_get(data, 'discovery_source'),
        'Additional Sources': list_to_str(data.get('additional_sources', [])),
        'Independent Source Count': safe_get(data, 'independent_source_count', default=0),
        'Electrical Source': safe_get(data, 'electrical_source'),
        'Water Type': safe_get(data, 'water_type'),
        'Legal Outcome': safe_get(data, 'legal_outcome'),
        'Project Refs': list_to_str(data.get('project_refs', [])),
    })
    return row

def build_incident_row_large(data):
    """All available fields."""
    row = build_incident_row_medium(data)
    row.update({
        'Date Precision': safe_get(data, 'date_precision'),
        'Electrical Source Detail': safe_get(data, 'electrical_source_detail'),
        'Voltage': safe_get(data, 'voltage'),
        'Fault Description': safe_get(data, 'fault_description'),
        'Latitude': safe_get(data, 'coordinates', 'lat'),
        'Longitude': safe_get(data, 'coordinates', 'lon'),
        'ESDPA Entry Numbers': list_to_str(data.get('esdpa_entry_numbers', [])),
        'ESDPA Date Listed': safe_get(data, 'esdpa_date_listed'),
        'ESDPA Data Issues': list_to_str(data.get('esdpa_data_issues', [])),
        'Research Notes': safe_get(data, 'research_notes'),
        'All Source URLs': get_all_source_urls(data),
        'All Sources Detail': get_all_sources_detail(data),
    })
    return row

def build_victim_rows(data):
    """Build one row per victim, with incident context."""
    victims = data.get('victims', []) or []
    if not victims or not isinstance(victims, list):
        # Still return one row for the incident even if no victims listed
        return [{
            'Incident ID': safe_get(data, 'incident_id'),
            'Date': safe_get(data, 'date'),
            'Year': safe_get(data, 'year'),
            'First Name': '',
            'Last Name': '',
            'Age': '',
            'Gender': '',
            'Race': '',
            'Outcome': safe_get(data, 'incident_type'),
            'Role': '',
            'Cause of Death': '',
            'State': safe_get(data, 'state'),
            'County': safe_get(data, 'county'),
            'City': safe_get(data, 'city'),
            'Body of Water': safe_get(data, 'body_of_water'),
            'Facility Name': safe_get(data, 'facility_name'),
            'Fatalities': safe_get(data, 'fatality_count', default=0),
            'Injuries': safe_get(data, 'injury_count', default=0),
            'Near Misses': safe_get(data, 'near_miss_count', default=0),
            'Verification Level': safe_get(data, 'verification_level'),
            'Notes': safe_get(data, 'notes'),
            'Source URL': get_first_source_url(data),
        }]

    rows = []
    for v in victims:
        if not isinstance(v, dict):
            continue
        rows.append({
            'Incident ID': safe_get(data, 'incident_id'),
            'Date': safe_get(data, 'date'),
            'Year': safe_get(data, 'year'),
            'First Name': safe_get(v, 'first_name'),
            'Last Name': safe_get(v, 'last_name'),
            'Age': safe_get(v, 'age'),
            'Gender': safe_get(v, 'gender'),
            'Race': safe_get(v, 'race'),
            'Outcome': safe_get(v, 'outcome'),
            'Role': safe_get(v, 'role'),
            'Cause of Death': safe_get(v, 'cause_of_death'),
            'State': safe_get(data, 'state'),
            'County': safe_get(data, 'county'),
            'City': safe_get(data, 'city'),
            'Body of Water': safe_get(data, 'body_of_water'),
            'Facility Name': safe_get(data, 'facility_name'),
            'Fatalities': safe_get(data, 'fatality_count', default=0),
            'Injuries': safe_get(data, 'injury_count', default=0),
            'Near Misses': safe_get(data, 'near_miss_count', default=0),
            'Verification Level': safe_get(data, 'verification_level'),
            'Notes': safe_get(data, 'notes'),
            'Source URL': get_first_source_url(data),
        })
    return rows

def style_worksheet(ws, header_fill_color="4472C4"):
    """Apply basic styling to a worksheet."""
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Auto-width columns (approximate)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                val_len = len(str(cell.value))
                # Cap at 60 for notes/URL columns
                max_len = max(max_len, min(val_len, 60))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Add filters
    ws.auto_filter.ref = ws.dimensions

def write_csv(rows, filepath, columns):
    """Write rows to CSV."""
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

def write_xlsx(incident_rows, victim_rows, filepath, incident_cols, victim_cols,
               sheet1_name="Incidents", sheet2_name="Victims", color="4472C4"):
    """Write XLSX with two sheets."""
    wb = Workbook()

    # Incidents sheet
    ws1 = wb.active
    ws1.title = sheet1_name
    ws1.append(incident_cols)
    for row in incident_rows:
        ws1.append([row.get(c, '') for c in incident_cols])
    style_worksheet(ws1, color)

    # Victims sheet
    ws2 = wb.create_sheet(title=sheet2_name)
    ws2.append(victim_cols)
    for row in victim_rows:
        ws2.append([row.get(c, '') for c in victim_cols])
    style_worksheet(ws2, color)

    wb.save(filepath)

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Parse all dataset files
    files = sorted(glob.glob(os.path.join(DATASET_DIR, "ESD-*.md")))
    print(f"Found {len(files)} dataset files")

    all_data = []
    errors = []
    for fp in files:
        data = parse_yaml_frontmatter(fp)
        if data:
            all_data.append(data)
        else:
            errors.append(os.path.basename(fp))

    print(f"Parsed {len(all_data)} files successfully, {len(errors)} errors")
    if errors:
        print(f"  Errors: {errors[:10]}")

    # Sort by date
    def sort_key(d):
        date = d.get('date', '0000-00-00')
        if not date: date = '0000-00-00'
        return str(date)
    all_data.sort(key=sort_key)

    # Build rows for each size
    small_incident_rows = [build_incident_row_small(d) for d in all_data]
    medium_incident_rows = [build_incident_row_medium(d) for d in all_data]
    large_incident_rows = [build_incident_row_large(d) for d in all_data]

    # Build victim rows
    victim_rows = []
    for d in all_data:
        victim_rows.extend(build_victim_rows(d))

    print(f"Incident rows: {len(small_incident_rows)}")
    print(f"Victim rows: {len(victim_rows)}")

    # Column definitions
    small_cols = list(small_incident_rows[0].keys())
    medium_cols = list(medium_incident_rows[0].keys())
    large_cols = list(large_incident_rows[0].keys())
    victim_cols = list(victim_rows[0].keys())

    # --- SMALL ---
    prefix = f"ESD-dataset-{DATE_STAMP}--small"
    write_csv(small_incident_rows, os.path.join(OUTPUT_DIR, f"{prefix}-incidents.csv"), small_cols)
    write_csv(victim_rows, os.path.join(OUTPUT_DIR, f"{prefix}-victims.csv"), victim_cols)
    write_xlsx(small_incident_rows, victim_rows,
               os.path.join(OUTPUT_DIR, f"{prefix}.xlsx"),
               small_cols, victim_cols, color="4472C4")
    print(f"Wrote {prefix}")

    # --- MEDIUM ---
    prefix = f"ESD-dataset-{DATE_STAMP}--medium"
    write_csv(medium_incident_rows, os.path.join(OUTPUT_DIR, f"{prefix}-incidents.csv"), medium_cols)
    write_xlsx(medium_incident_rows, victim_rows,
               os.path.join(OUTPUT_DIR, f"{prefix}.xlsx"),
               medium_cols, victim_cols, color="2E75B6")
    print(f"Wrote {prefix}")

    # --- LARGE ---
    prefix = f"ESD-dataset-{DATE_STAMP}--large"
    write_csv(large_incident_rows, os.path.join(OUTPUT_DIR, f"{prefix}-incidents.csv"), large_cols)
    write_xlsx(large_incident_rows, victim_rows,
               os.path.join(OUTPUT_DIR, f"{prefix}.xlsx"),
               large_cols, victim_cols, color="1F4E79")
    print(f"Wrote {prefix}")

    print(f"\nAll files written to {OUTPUT_DIR}")

if __name__ == "__main__":
    main()
